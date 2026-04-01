"""Build Excel report with Shadow Market Summary and Block Group Detail sheets."""

from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from market_config import MarketConfig

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(size=10, color="666666")
THIN_BORDER_TOP = Border(top=Side(style="thin"))
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
SECTION_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)

NUM_FMT = "#,##0"
DEC_FMT = "#,##0.0"
PCT_FMT = "0.0%"
DIST_FMT = "0.00"
OCC_FMT = "0.00"
CHANGE_FMT = "+#,##0;-#,##0;0"
CHANGE_DEC_FMT = "+#,##0.0;-#,##0.0;0.0"
PP_FMT = "+0.0;-0.0;0.0"


def _write_header_row(ws, row, headers, widths):
    """Write a styled header row."""
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_agg_row(ws, row, label, agg, bold=False):
    """Write one ring-aggregation row on the summary sheet."""
    font = BOLD if bold else None
    border = THIN_BORDER_TOP if bold else None

    values = [
        label,
        agg["block_groups"],
        agg["total_units"],
        agg["renter_units"],
        agg["renter_units_sub50"],
        agg["renter_units_50plus"],
        agg["sub50_ratio"] / 100,
        agg["renter_15_24"],
        agg["renter_25_34"],
        agg["pct_15_24_of_renter"] / 100,
        agg["shadow_hhs"],
        agg["shadow_pop"],
        agg["avg_occ_sub50"],
    ]
    fmts = [None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, PCT_FMT,
            NUM_FMT, NUM_FMT, PCT_FMT, DEC_FMT, DEC_FMT, OCC_FMT]

    for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        if fmt:
            cell.number_format = fmt
        if font:
            cell.font = font
        if border:
            cell.border = border
        if col_idx > 1:
            cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Per-year summary sheet
# ---------------------------------------------------------------------------

SUMMARY_HEADERS = [
    "Ring", "Block\nGroups", "Total\nUnits", "Renter\nUnits", "Renter\nUnits\nSub-50",
    "Renter\nUnits\n50+", "Sub-50\nRatio", "Renters\n15-24", "Renters\n25-34",
    "% 15-24\n(of Renter)", "Shadow Mkt\nUnits\n(15-24)", "Shadow Mkt\nPop\n(15-24)", "Avg Occ\nSub-50",
]
SUMMARY_WIDTHS = [12, 10, 12, 12, 12, 10, 10, 12, 12, 13, 12, 12, 10]


def _build_summary(wb, result, year: int, config: MarketConfig, sheet_name: str | None = None):
    if sheet_name:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb.active
        ws.title = "Shadow Market Summary"

    campus_list = ", ".join(config.campuses.keys())
    ncols = len(SUMMARY_HEADERS)

    ws.cell(row=1, column=1, value=f"Shadow Market Analysis \u2014 {config.name}").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=2, column=1, value=f"Census ACS 5-Year ({year})  \u2022  Generated {date.today().isoformat()}").font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    _write_header_row(ws, 4, SUMMARY_HEADERS, SUMMARY_WIDTHS)

    row = 5
    for label in config.ring_labels:
        _write_agg_row(ws, row, label, result["rings"][label])
        row += 1

    _write_agg_row(ws, row, "TOTAL", result["total"], bold=True)
    row += 2

    county_list = " + ".join(f"{n} County" for n in config.county_names)
    notes = [
        "Methodology Notes:",
        f"\u2022 Data source: U.S. Census Bureau, ACS 5-Year Estimates ({year}), Tables B25007, B25032, B25033",
        f"\u2022 Distance calculated from nearest {config.name} campus ({campus_list}) to block group centroid",
        "\u2022 Age cohort is 15-24 (finest Census granularity); closely approximates traditional college-age renters",
        "\u2022 Sub-50 = renter units in buildings with <50 units (B25032); excludes 50+ unit buildings already tracked internally",
        "\u2022 Shadow Mkt Units (15-24) = Renters 15-24 \u00d7 Sub-50 Ratio (per block group)",
        "\u2022 Shadow Mkt Pop = Shadow Mkt HHs \u00d7 Avg Occupancy per Sub-50 Unit (B25033 pop allocated by B25032 unit ratio)",
        f"\u2022 Block groups drawn from {county_list}; assigned to nearest campus \u2014 no double-counting",
    ]
    for note in notes:
        cell = ws.cell(row=row, column=1, value=note)
        if note.startswith("Methodology"):
            cell.font = BOLD
        else:
            cell.font = Font(size=9, color="555555")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        row += 1


# ---------------------------------------------------------------------------
# Per-year detail sheet
# ---------------------------------------------------------------------------

def _build_detail(wb, detail, config: MarketConfig, sheet_name: str = "Block Group Detail"):
    ws = wb.create_sheet(sheet_name)

    headers = [
        "GEOID", "Name", "Nearest Campus", "Ring", "Distance (mi)",
        "Renter\nUnits", "Renter\nUnits\nSub-50", "Renter\nUnits\n50+", "Sub-50\nRatio",
        "Renter\nPop\n(B25033)", "Renter Pop\nSub-50\n(Est.)", "Avg Occ\nper Sub-50\nUnit",
        "Renters\n15-24", "Shadow Mkt\nUnits\n(15-24)", "Shadow Mkt\nPop\n(15-24)",
    ]
    widths = [16, 50, 16, 12, 11, 11, 11, 10, 10, 12, 12, 12, 11, 12, 12]
    _write_header_row(ws, 1, headers, widths)

    for i, rec in enumerate(detail):
        row = i + 2
        values = [
            rec["geoid"],
            rec["name"],
            rec["nearest_campus"],
            rec["ring"],
            rec["distance_mi"],
            rec["renter_total"],
            rec["renter_units_sub50"],
            rec["renter_units_50plus"],
            rec["sub50_ratio"],
            rec["renter_pop"],
            rec["renter_pop_sub50"],
            rec["avg_occ_sub50"],
            rec["renter_15_24"],
            rec["shadow_hhs"],
            rec["shadow_pop"],
        ]
        fmts = [None, None, None, None, DIST_FMT,
                NUM_FMT, NUM_FMT, NUM_FMT, PCT_FMT,
                NUM_FMT, DEC_FMT, OCC_FMT,
                NUM_FMT, DEC_FMT, DEC_FMT]

        for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if fmt:
                cell.number_format = fmt
            if col_idx >= 5:
                cell.alignment = Alignment(horizontal="center")

        if i % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = ALT_FILL

    last_row = len(detail) + 1
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Master comparison sheet
# ---------------------------------------------------------------------------

def _write_metric_section(ws, row, section_label, yearly_results, years, metrics, agg_key):
    """Write a section (ring or TOTAL) of the master comparison table."""
    num_headers = 2 + len(years) + 1

    # Section header row
    for col_idx in range(1, num_headers + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = SECTION_FILL
    ws.cell(row=row, column=1, value=section_label).font = SECTION_FONT
    if section_label == "TOTAL":
        ws.cell(row=row, column=1).border = THIN_BORDER_TOP
    row += 1

    for metric_label, metric_key, val_fmt, change_fmt in metrics:
        ws.cell(row=row, column=2, value=metric_label)

        is_pct = val_fmt == PCT_FMT
        vals = []
        for ci, y in enumerate(years):
            agg = yearly_results[y]["total"] if agg_key == "total" else yearly_results[y]["rings"][agg_key]
            v = agg[metric_key]
            if is_pct:
                v = v / 100
            vals.append(v)
            cell = ws.cell(row=row, column=3 + ci, value=v)
            cell.number_format = val_fmt
            cell.alignment = Alignment(horizontal="center")

        change = vals[-1] - vals[0]
        if is_pct:
            change = change * 100
        change_cell = ws.cell(row=row, column=3 + len(years), value=change)
        change_cell.number_format = change_fmt
        change_cell.alignment = Alignment(horizontal="center")
        if is_pct:
            change_cell.number_format = '+0.0"pp";-0.0"pp";0.0"pp"'

        row += 1

    return row + 1


def _build_master(wb, yearly_results: dict[int, dict], config: MarketConfig):
    """Build the master trend comparison as the first sheet."""
    ws = wb.active
    ws.title = "Master Comparison"

    years = sorted(yearly_results.keys())
    first_year, last_year = years[0], years[-1]
    year_range = f"{first_year}\u2013{last_year}"

    ws.cell(row=1, column=1, value=f"Shadow Market Trends \u2014 {config.name}").font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.cell(row=2, column=1, value=f"ACS 5-Year Comparison ({year_range})  \u2022  Generated {date.today().isoformat()}").font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    headers = ["Ring", "Metric"] + [str(y) for y in years] + [f"Change ({str(first_year)[2:]}\u2192{str(last_year)[2:]})"]
    widths = [12, 22] + [14] * len(years) + [18]
    _write_header_row(ws, 4, headers, widths)

    # All metrics including shadow market
    unit_metrics = [
        ("Block Groups", "block_groups", NUM_FMT, CHANGE_FMT),
        ("Total Units", "total_units", NUM_FMT, CHANGE_FMT),
        ("Renter Units", "renter_units", NUM_FMT, CHANGE_FMT),
        ("Renter Units Sub-50", "renter_units_sub50", NUM_FMT, CHANGE_FMT),
        ("Renter Units 50+", "renter_units_50plus", NUM_FMT, CHANGE_FMT),
        ("Sub-50 Ratio", "sub50_ratio", PCT_FMT, PP_FMT),
        ("Renters 15-24", "renter_15_24", NUM_FMT, CHANGE_FMT),
        ("Renters 25-34", "renter_25_34", NUM_FMT, CHANGE_FMT),
        ("% 15-24 (of Renter)", "pct_15_24_of_renter", PCT_FMT, PP_FMT),
        ("Avg Occ per Sub-50 Unit", "avg_occ_sub50", OCC_FMT, "+0.00;-0.00;0.00"),
        ("Shadow Market HHs", "shadow_hhs", DEC_FMT, CHANGE_DEC_FMT),
        ("Shadow Market Pop", "shadow_pop", DEC_FMT, CHANGE_DEC_FMT),
    ]

    row = 5

    # --- SHADOW MARKET POPULATION section (amber) ---
    pop_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    pop_header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = pop_header_fill
    ws.cell(row=row, column=1, value="SHADOW MARKET POPULATION").font = Font(bold=True, color="FFFFFF", size=11)
    ws.cell(row=row, column=2, value="(per-BG occupancy from B25032/B25033)").font = Font(italic=True, color="FFFFFF", size=9)
    row += 1

    sm_metrics = [
        ("Shadow Mkt Units (15-24)", "shadow_hhs"),
        ("Shadow Mkt Pop (15-24)", "shadow_pop"),
        ("Avg Occ per Sub-50 Unit", "avg_occ_sub50"),
    ]
    sm_sections = [(label, label) for label in config.ring_labels] + [("TOTAL", "total")]

    for section_label, agg_key in sm_sections:
        is_total = agg_key == "total"
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).fill = pop_fill
        cell_a = ws.cell(row=row, column=1, value=section_label)
        cell_a.font = BOLD
        cell_a.fill = pop_fill
        if is_total:
            cell_a.border = THIN_BORDER_TOP
        row += 1

        for metric_label, metric_key in sm_metrics:
            ws.cell(row=row, column=2, value=metric_label).fill = pop_fill
            ws.cell(row=row, column=1).fill = pop_fill
            is_occ = metric_key == "avg_occ_sub50"
            vals = []
            for ci, y in enumerate(years):
                agg = yearly_results[y]["total"] if is_total else yearly_results[y]["rings"][agg_key]
                v = agg[metric_key]
                vals.append(v)
                cell = ws.cell(row=row, column=3 + ci, value=v)
                cell.number_format = OCC_FMT if is_occ else DEC_FMT
                cell.alignment = Alignment(horizontal="center")
                cell.fill = pop_fill
            change = vals[-1] - vals[0]
            change_cell = ws.cell(row=row, column=3 + len(years), value=change)
            change_cell.number_format = "+0.00;-0.00;0.00" if is_occ else CHANGE_DEC_FMT
            change_cell.alignment = Alignment(horizontal="center")
            change_cell.fill = pop_fill
            if is_total:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).font = BOLD
            row += 1

    row += 1  # blank row

    # --- TOTAL section ---
    row = _write_metric_section(ws, row, "TOTAL", yearly_results, years, unit_metrics, "total")

    # --- Per-ring sections ---
    for ring_label in config.ring_labels:
        row = _write_metric_section(ws, row, ring_label, yearly_results, years, unit_metrics, ring_label)

    # Notes
    row += 1
    notes = [
        "Notes:",
        "\u2022 Shadow Market HHs = Renters 15-24 \u00d7 Sub-50 Ratio (per block group)",
        "\u2022 Shadow Mkt Pop (15-24) = Shadow Mkt Units \u00d7 Avg Occupancy per Sub-50 Unit (per block group)",
        "\u2022 Sub-50 ratio and avg occupancy are calculated per block group, then aggregated (not a flat multiplier)",
        "\u2022 Sub-50 = buildings with <50 units (B25032); 50+ unit buildings excluded to avoid overlap with internal data",
        "\u2022 Block group boundaries are from the 2020 Census; ACS estimates vary by vintage",
        f'\u2022 "Change" column shows absolute difference ({first_year} \u2192 {last_year}); "pp" = percentage points',
    ]
    for note in notes:
        cell = ws.cell(row=row, column=1, value=note)
        if note.startswith("Notes"):
            cell.font = BOLD
        else:
            cell.font = Font(size=9, color="555555")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1


# ---------------------------------------------------------------------------
# Public builders
# ---------------------------------------------------------------------------

def build_report(result: dict, year: int, config: MarketConfig) -> bytes:
    """Create single-year workbook and return as bytes."""
    wb = Workbook()
    _build_summary(wb, result, year, config)
    _build_detail(wb, result["detail"], config)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_master_report(yearly_results: dict[int, dict], config: MarketConfig) -> bytes:
    """Create multi-year master workbook and return as bytes."""
    wb = Workbook()

    # Master comparison sheet (uses wb.active)
    _build_master(wb, yearly_results, config)

    # Per-year summary + detail sheets
    for year in sorted(yearly_results.keys()):
        result = yearly_results[year]
        _build_summary(wb, result, year, config, sheet_name=f"{year} Summary")
        _build_detail(wb, result["detail"], config, sheet_name=f"{year} Detail")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
