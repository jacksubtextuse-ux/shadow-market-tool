"""Build Excel reports from CoStar building-level data."""

from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from market_config import MarketConfig

# Reuse styles from report.py
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(size=10, color="666666")
THIN_BORDER_TOP = Border(top=Side(style="thin"))
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

NUM_FMT = "#,##0"
DEC_FMT = "#,##0.0"
OCC_FMT = "0.00"
DIST_FMT = "0.00"


def _write_header_row(ws, row, headers, widths):
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_costar_beds_report(result: dict, config: MarketConfig) -> bytes:
    """Build Excel report with beds-only CoStar analysis (no occupancy)."""
    wb = Workbook()
    _build_beds_summary(wb, result, config)
    _build_beds_detail(wb, result["detail"], config)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_costar_occupancy_report(result: dict, config: MarketConfig) -> bytes:
    """Build Excel report with CoStar + Census occupancy analysis."""
    wb = Workbook()
    _build_occ_summary(wb, result, config)
    _build_occ_detail(wb, result["detail"], config)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_costar_combined_report(result: dict, config: MarketConfig, include_graduates: bool) -> bytes:
    """Build Excel report with combined CoStar + Census 2-4 + college-age analysis."""
    wb = Workbook()
    _build_combined_summary(wb, result, config, include_graduates)
    _build_combined_detail(wb, result["detail"], config, include_graduates)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- Combined summary ----

def _build_combined_summary(wb, result, config, include_graduates):
    ws = wb.active
    age_label = "18-24" if include_graduates else "18-21"
    ws.title = "Shadow Market Summary"

    ws.cell(row=1, column=1, value=f"{config.name} — Shadow Market Analysis").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"CoStar (5-49 units) + Census (2-4 units) | College age: {age_label} | Generated {date.today()}"
    ).font = SUBTITLE_FONT

    headers = ["Ring", "CoStar\nBuildings", "CoStar\nUnits", "CoStar\nBeds",
               "Census\n2-4 BGs", "Census\n2-4 Units",
               "Total\nUnits", "Avg Occ", "Est.\nPop",
               f"College %\n({age_label})", "Shadow\nPop"]
    widths = [12, 12, 12, 12, 12, 12, 12, 10, 12, 12, 12]
    _write_header_row(ws, 4, headers, widths)

    row = 5
    for label in config.ring_labels:
        agg = result["rings"][label]
        _write_combined_row(ws, row, label, agg)
        row += 1

    _write_combined_row(ws, row, "TOTAL", result["total"], bold=True)

    # Methodology
    note_row = row + 2
    NOTE_FONT = Font(size=9, color="555555")
    NOTE_BOLD = Font(size=9, color="333333", bold=True)

    ws.cell(row=note_row, column=1, value="METHODOLOGY").font = NOTE_BOLD
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)

    notes = [
        "1. CoStar buildings: sub-50 unit multi-family properties (excludes Student type + Demolished).",
        "2. Census 2-4 units: renter-occupied 1-unit and 2-4 unit structures from ACS B25032 per block group.",
        "   CoStar does not capture these small properties — Census fills the gap.",
        "3. Total Units = CoStar units + Census 2-4 units per ring.",
        "4. Avg Occ = Census avg occupancy per sub-50 unit (B25033 pop / B25032 units) per block group.",
        "5. Est. Pop = units × avg occupancy, computed per block group then summed by ring.",
        f"6. Renter 15-24 share = (renters 15-24 / total renters) from Census B25007, per block group.",
        f"7. Tightening ratio = (pop {age_label}) / (pop 15-24) from Census B01001, per block group.",
        f"   This narrows the 15-24 renter count to actual college-age ({age_label}).",
        f"8. College renter % = Renter 15-24 share × Tightening ratio, per block group.",
        "9. Shadow Pop = Est. Pop × College renter % — estimated college-age renter population.",
        "10. All percentages and occupancy rates are per individual block group, NOT aggregated averages.",
    ]
    for i, note in enumerate(notes):
        r = note_row + 1 + i
        ws.cell(row=r, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)


def _write_combined_row(ws, row, label, agg, bold=False):
    font = BOLD if bold else None
    border = THIN_BORDER_TOP if bold else None

    total_units = agg.get("total_units", 0)
    est_pop = agg.get("est_pop", 0)
    shadow_pop = agg.get("shadow_pop", 0)
    college_pct = (shadow_pop / est_pop) if est_pop > 0 else 0

    values = [label, agg.get("buildings", 0), agg.get("costar_units", 0), agg.get("costar_beds", 0),
              agg.get("census_2to4_bgs", 0), agg.get("census_2to4_units", 0),
              total_units, agg.get("avg_occ", 0), round(est_pop, 0),
              college_pct, round(shadow_pop, 0)]
    fmts = [None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT,
            NUM_FMT, OCC_FMT, NUM_FMT, "0.0%", NUM_FMT]

    for col, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell = ws.cell(row=row, column=col, value=val)
        if fmt:
            cell.number_format = fmt
        if font:
            cell.font = font
        if border:
            cell.border = border


# ---- Combined detail ----

def _build_combined_detail(wb, detail, config, include_graduates):
    ws = wb.create_sheet("Building Detail")
    age_label = "18-24" if include_graduates else "18-21"

    headers = ["Source", "Property / BG Name", "Address / GEOID", "Nearest Campus", "Ring",
               "Distance (mi)", "Units", "Beds", "Avg Occ", f"College %\n({age_label})",
               "Est. Pop", "Shadow Pop"]
    widths = [12, 28, 28, 18, 12, 12, 10, 10, 10, 12, 12, 12]
    _write_header_row(ws, 1, headers, widths)

    for i, rec in enumerate(detail, start=2):
        ws.cell(row=i, column=1, value=rec.get("source", ""))
        ws.cell(row=i, column=2, value=rec.get("name", ""))
        ws.cell(row=i, column=3, value=rec.get("address", ""))
        ws.cell(row=i, column=4, value=rec.get("nearest_campus", ""))
        ws.cell(row=i, column=5, value=rec.get("ring", ""))
        ws.cell(row=i, column=6, value=rec.get("distance_mi", 0)).number_format = DIST_FMT
        ws.cell(row=i, column=7, value=rec.get("units", 0)).number_format = NUM_FMT
        ws.cell(row=i, column=8, value=rec.get("beds", 0)).number_format = NUM_FMT
        ws.cell(row=i, column=9, value=rec.get("census_avg_occ", 0)).number_format = OCC_FMT
        ws.cell(row=i, column=10, value=(rec.get("college_pct", 0) / 100)).number_format = "0.0%"
        ws.cell(row=i, column=11, value=rec.get("est_pop", 0)).number_format = DEC_FMT
        ws.cell(row=i, column=12, value=rec.get("shadow_pop", 0)).number_format = DEC_FMT

        if i % 2 == 0:
            for col in range(1, 13):
                ws.cell(row=i, column=col).fill = ALT_FILL

    ws.auto_filter.ref = f"A1:L{len(detail) + 1}"
    ws.freeze_panes = "A2"


# ---- Beds-only summary ----

def _build_beds_summary(wb, result, config):
    ws = wb.active
    ws.title = "CoStar Summary"

    ws.cell(row=1, column=1, value=f"{config.name} — CoStar Building Analysis").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Sub-50 unit buildings within {max(config.ring_miles)} mi | Generated {date.today()}").font = SUBTITLE_FONT

    headers = ["Ring", "Buildings", "Units", "Beds", "Studios", "1 BR", "2 BR", "3 BR", "4 BR"]
    widths = [12, 12, 12, 12, 10, 10, 10, 10, 10]
    _write_header_row(ws, 4, headers, widths)

    row = 5
    for label in config.ring_labels:
        agg = result["rings"][label]
        _write_summary_row(ws, row, label, agg)
        row += 1

    # Total row
    _write_summary_row(ws, row, "TOTAL", result["total"], bold=True)

    # Methodology notes
    note_row = row + 2
    NOTE_FONT = Font(size=9, color="555555")
    NOTE_BOLD = Font(size=9, color="333333", bold=True)

    ws.cell(row=note_row, column=1, value="METHODOLOGY").font = NOTE_BOLD
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)

    notes = [
        "1. Source: CoStar building-level data filtered to sub-50 unit properties (50+ excluded — tracked internally).",
        "2. Demolished properties are excluded from the analysis.",
        "3. Beds = total bed count per building (closer proxy to population than unit count).",
        "4. No Census occupancy applied — this is raw CoStar data only.",
        "5. Distance rings measured from campus center using haversine (great-circle) distance.",
    ]
    for i, note in enumerate(notes):
        r = note_row + 1 + i
        ws.cell(row=r, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)


def _write_summary_row(ws, row, label, agg, bold=False):
    font = BOLD if bold else None
    border = THIN_BORDER_TOP if bold else None

    values = [label, agg["buildings"], agg["units"], agg["beds"],
              agg["studio"], agg["br1"], agg["br2"], agg["br3"], agg["br4"]]
    fmts = [None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT]

    for col, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell = ws.cell(row=row, column=col, value=val)
        if fmt:
            cell.number_format = fmt
        if font:
            cell.font = font
        if border:
            cell.border = border


# ---- Beds-only detail ----

def _build_beds_detail(wb, detail, config):
    ws = wb.create_sheet("Building Detail")

    headers = ["Property Name", "Address", "Nearest Campus", "Ring", "Distance (mi)",
               "Units", "Beds", "Studios", "1 BR", "2 BR", "3 BR", "4 BR", "Year Built"]
    widths = [28, 32, 18, 12, 12, 10, 10, 10, 10, 10, 10, 10, 12]
    _write_header_row(ws, 1, headers, widths)

    for i, rec in enumerate(detail, start=2):
        ws.cell(row=i, column=1, value=rec["name"])
        ws.cell(row=i, column=2, value=rec["address"])
        ws.cell(row=i, column=3, value=rec["nearest_campus"])
        ws.cell(row=i, column=4, value=rec["ring"])
        ws.cell(row=i, column=5, value=rec["distance_mi"]).number_format = DIST_FMT
        ws.cell(row=i, column=6, value=rec["units"]).number_format = NUM_FMT
        ws.cell(row=i, column=7, value=rec["beds"]).number_format = NUM_FMT
        ws.cell(row=i, column=8, value=rec["studio"]).number_format = NUM_FMT
        ws.cell(row=i, column=9, value=rec["br1"]).number_format = NUM_FMT
        ws.cell(row=i, column=10, value=rec["br2"]).number_format = NUM_FMT
        ws.cell(row=i, column=11, value=rec["br3"]).number_format = NUM_FMT
        ws.cell(row=i, column=12, value=rec["br4"]).number_format = NUM_FMT
        ws.cell(row=i, column=13, value=rec["year_built"] or "").number_format = "0"

        if i % 2 == 0:
            for col in range(1, 14):
                ws.cell(row=i, column=col).fill = ALT_FILL

    ws.auto_filter.ref = f"A1:M{len(detail) + 1}"
    ws.freeze_panes = "A2"


# ---- Occupancy summary ----

def _build_occ_summary(wb, result, config):
    ws = wb.active
    ws.title = "CoStar + Census"

    ws.cell(row=1, column=1, value=f"{config.name} — CoStar + Census Shadow Market").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Sub-50 unit buildings x Census 15-24 renter share x occupancy | Generated {date.today()}").font = SUBTITLE_FONT

    headers = ["Ring", "Buildings", "Units", "Beds", "Studios", "1 BR", "2 BR", "3 BR", "4 BR",
               "Avg Occ", "Est. Pop",
               "Shadow Mkt\nUnits (15-24)", "Shadow Mkt\nPop (15-24)"]
    widths = [12, 12, 12, 12, 10, 10, 10, 10, 10, 10, 12, 16, 16]
    _write_header_row(ws, 4, headers, widths)

    row = 5
    for label in config.ring_labels:
        agg = result["rings"][label]
        _write_occ_summary_row(ws, row, label, agg)
        row += 1

    _write_occ_summary_row(ws, row, "TOTAL", result["total"], bold=True)

    # Methodology notes
    note_row = row + 2
    NOTE_FONT = Font(size=9, color="555555")
    NOTE_BOLD = Font(size=9, color="333333", bold=True)

    ws.cell(row=note_row, column=1, value="METHODOLOGY").font = NOTE_BOLD
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)

    notes = [
        "1. Source: CoStar building-level data filtered to sub-50 unit properties (50+ excluded — tracked internally).",
        "2. Demolished properties are excluded from the analysis.",
        "3. Each building is matched to the nearest Census block group centroid by geographic distance.",
        "4. Census Avg Occ = average people per sub-50 unit in that block group, derived from ACS tables B25032 + B25033.",
        "   Formula: (pop_1unit + pop_2to4 + pop_5plus × sub50_share) / renter_units_sub50",
        "5. Est. Pop = Building Units × Census Avg Occ from the matched block group.",
        "6. Shadow Mkt Units (15-24) = Building Units × (BG Renters 15-24 / BG Total Renters) from Census B25007.",
        "7. Shadow Mkt Pop (15-24) = Shadow Mkt Units × Census Avg Occ.",
        "8. Avg Occ in summary = weighted average across all buildings in that ring (total est. pop / total units).",
        "9. Distance rings measured from campus center using haversine (great-circle) distance.",
    ]
    for i, note in enumerate(notes):
        r = note_row + 1 + i
        ws.cell(row=r, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)


def _write_occ_summary_row(ws, row, label, agg, bold=False):
    font = BOLD if bold else None
    border = THIN_BORDER_TOP if bold else None

    values = [label, agg["buildings"], agg["units"], agg["beds"],
              agg["studio"], agg["br1"], agg["br2"], agg["br3"], agg["br4"],
              agg.get("avg_occ", 0), agg.get("est_pop", 0),
              agg.get("shadow_units_15_24", 0), agg.get("shadow_pop_15_24", 0)]
    fmts = [None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT,
            OCC_FMT, NUM_FMT, DEC_FMT, DEC_FMT]

    for col, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell = ws.cell(row=row, column=col, value=val)
        if fmt:
            cell.number_format = fmt
        if font:
            cell.font = font
        if border:
            cell.border = border


# ---- Occupancy detail ----

def _build_occ_detail(wb, detail, config):
    ws = wb.create_sheet("Building Detail")

    headers = ["Property Name", "Address", "Nearest Campus", "Ring", "Distance (mi)",
               "Units", "Beds", "Studios", "1 BR", "2 BR", "3 BR", "4 BR", "Year Built",
               "Matched BG", "Census Avg Occ", "% Renters 15-24", "Est. Pop",
               "Shadow Units (15-24)", "Shadow Pop (15-24)"]
    widths = [28, 32, 18, 12, 12, 10, 10, 10, 10, 10, 10, 10, 12, 16, 14, 14, 12, 16, 16]
    _write_header_row(ws, 1, headers, widths)

    for i, rec in enumerate(detail, start=2):
        ws.cell(row=i, column=1, value=rec["name"])
        ws.cell(row=i, column=2, value=rec["address"])
        ws.cell(row=i, column=3, value=rec["nearest_campus"])
        ws.cell(row=i, column=4, value=rec["ring"])
        ws.cell(row=i, column=5, value=rec["distance_mi"]).number_format = DIST_FMT
        ws.cell(row=i, column=6, value=rec["units"]).number_format = NUM_FMT
        ws.cell(row=i, column=7, value=rec["beds"]).number_format = NUM_FMT
        ws.cell(row=i, column=8, value=rec["studio"]).number_format = NUM_FMT
        ws.cell(row=i, column=9, value=rec["br1"]).number_format = NUM_FMT
        ws.cell(row=i, column=10, value=rec["br2"]).number_format = NUM_FMT
        ws.cell(row=i, column=11, value=rec["br3"]).number_format = NUM_FMT
        ws.cell(row=i, column=12, value=rec["br4"]).number_format = NUM_FMT
        ws.cell(row=i, column=13, value=rec["year_built"] or "").number_format = "0"
        ws.cell(row=i, column=14, value=rec.get("matched_bg", ""))
        ws.cell(row=i, column=15, value=rec.get("census_avg_occ", 0)).number_format = OCC_FMT
        ws.cell(row=i, column=16, value=(rec.get("pct_15_24", 0) / 100) if rec.get("pct_15_24") else 0).number_format = "0.0%"
        ws.cell(row=i, column=17, value=rec.get("est_pop", 0)).number_format = DEC_FMT
        ws.cell(row=i, column=18, value=rec.get("shadow_units_15_24", 0)).number_format = DEC_FMT
        ws.cell(row=i, column=19, value=rec.get("shadow_pop_15_24", 0)).number_format = DEC_FMT

        if i % 2 == 0:
            for col in range(1, 20):
                ws.cell(row=i, column=col).fill = ALT_FILL

    ws.auto_filter.ref = f"A1:S{len(detail) + 1}"
    ws.freeze_panes = "A2"
